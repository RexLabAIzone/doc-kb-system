import os
import csv
import logging
from collections import Counter, defaultdict

from config import setup_logging

logger = setup_logging("reporter")


def export_excel(results, output_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl is required for Excel export. Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    green_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    _build_summary_sheet(wb, results, header_fill, header_font, thin_border)
    _build_issues_sheet(wb, results, red_fill, yellow_fill, header_fill, header_font, thin_border)
    _build_all_files_sheet(wb, results, red_fill, yellow_fill, green_fill, header_fill, header_font, thin_border)

    wb.save(output_path)
    logger.info("Excel report saved to %s", output_path)


def _build_summary_sheet(wb, results, header_fill, header_font, thin_border):
    ws = wb.active
    ws.title = "Summary"

    total = len(results)
    status_counts = Counter(r["status"] for r in results)
    category_counts = Counter(r["category"] for r in results)
    ext_counts = Counter(r["ext"].lower() for r in results)

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    summary_data = [
        ("Total Files", total),
        ("OK", status_counts.get("ok", 0)),
        ("Warning", status_counts.get("warning", 0)),
        ("Error", status_counts.get("error", 0)),
        ("", ""),
        ("Files by Category", ""),
    ]
    for cat, cnt in sorted(category_counts.items(), key=lambda x: -x[1]):
        summary_data.append((f"  {cat}", cnt))

    summary_data.append(("", ""))
    summary_data.append(("Top 20 Extensions", ""))
    for ext, cnt in ext_counts.most_common(20):
        summary_data.append((f"  {ext}", cnt))

    for row_idx, (label, val) in enumerate(summary_data, 2):
        ws.cell(row=row_idx, column=1, value=str(label)).border = thin_border
        cell = ws.cell(row=row_idx, column=2, value=val)
        cell.border = thin_border
        if val == "Error" and isinstance(val, int) and val > 0:
            cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.freeze_panes = "A2"


def _build_issues_sheet(wb, results, red_fill, yellow_fill, header_fill, header_font, thin_border):
    ws = wb.create_sheet("Issues")
    headers = ["File Path", "Severity", "Type", "Message"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    for r in results:
        for issue in r.get("issues", []):
            ws.cell(row=row_idx, column=1, value=r["path"]).border = thin_border
            sev_cell = ws.cell(row=row_idx, column=2, value=issue.get("severity", ""))
            sev_cell.border = thin_border
            ws.cell(row=row_idx, column=3, value=issue.get("type", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=issue.get("message", "")).border = thin_border
            sev = issue.get("severity", "")
            if sev == "error":
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = red_fill
            elif sev == "warning":
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = yellow_fill
            row_idx += 1

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 70
    ws.freeze_panes = "A2"


def _build_all_files_sheet(wb, results, red_fill, yellow_fill, green_fill, header_fill, header_font, thin_border):
    ws = wb.create_sheet("All Files")
    headers = ["Path", "Name", "Extension", "Size (bytes)", "Category", "Status",
               "Encoding", "Charset Confidence", "Page Count",
               "Image Dimensions", "Duration (ms)", "Issues"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        vals = [
            r["path"], r["name"], r["ext"], r["size"], r["category"], r["status"],
            r.get("encoding", ""), r.get("charset_confidence", ""),
            r.get("page_count", ""),
            str(r.get("image_dimensions", "")) if r.get("image_dimensions") else "",
            f"{r.get('duration_ms', 0):.1f}",
            "; ".join(i.get("message", "") for i in r.get("issues", [])),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            col_letter = get_column_letter(col)
            max_len = max(len(str(val)), 8) if val else 8
            current_width = ws.column_dimensions[col_letter].width or 8
            ws.column_dimensions[col_letter].width = max(current_width, min(max_len + 2, 60))

        status = r["status"]
        if status == "error":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill
        elif status == "warning":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = yellow_fill
        else:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = green_fill

    ws.freeze_panes = "A2"


def export_csv(results, output_path):
    try:
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            headers = ["Path", "Name", "Extension", "Size (bytes)", "Category", "Status",
                       "Encoding", "Charset Confidence", "Page Count",
                       "Image Dimensions", "Duration (ms)", "Issues"]
            writer.writerow(headers)
            for r in results:
                row = [
                    r["path"], r["name"], r["ext"], r["size"], r["category"], r["status"],
                    r.get("encoding", ""), r.get("charset_confidence", ""),
                    r.get("page_count", ""),
                    str(r.get("image_dimensions", "")) if r.get("image_dimensions") else "",
                    f"{r.get('duration_ms', 0):.1f}",
                    "; ".join(i.get("message", "") for i in r.get("issues", [])),
                ]
                writer.writerow(row)
        logger.info("CSV report saved to %s", output_path)
    except Exception as e:
        logger.error("Failed to export CSV: %s", e)


def generate_summary_text(results):
    total = len(results)
    status_counts = Counter(r["status"] for r in results)
    category_counts = Counter(r["category"] for r in results)
    ext_counts = Counter(r["ext"].lower() for r in results)
    all_issues = [i for r in results for i in r.get("issues", [])]
    issue_type_counts = Counter(i["type"] for i in all_issues)
    issue_severity_counts = Counter(i["severity"] for i in all_issues)

    total_size = sum(r["size"] for r in results)
    total_time = sum(r.get("duration_ms", 0) for r in results)

    lines = []
    lines.append("=" * 70)
    lines.append("DOCUMENT HEALTH CHECK SUMMARY")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"Total files scanned: {total:,}")
    lines.append(f"Total size: {_format_size(total_size)}")
    lines.append(f"Total check time: {total_time / 1000:.1f}s")
    lines.append("")
    lines.append("STATUS BREAKDOWN:")
    lines.append(f"  OK:      {status_counts.get('ok', 0):,}")
    lines.append(f"  Warning: {status_counts.get('warning', 0):,}")
    lines.append(f"  Error:   {status_counts.get('error', 0):,}")
    lines.append("")
    ok_rate = (status_counts.get("ok", 0) / total * 100) if total > 0 else 0
    lines.append(f"Health rate: {ok_rate:.1f}%")
    lines.append("")
    lines.append("FILES BY CATEGORY:")
    for cat, cnt in sorted(category_counts.items(), key=lambda x: -x[1]):
        lines.append(f"  {cat}: {cnt:,}")
    lines.append("")
    lines.append("TOP 10 EXTENSIONS:")
    for ext, cnt in ext_counts.most_common(10):
        lines.append(f"  {ext}: {cnt:,}")
    lines.append("")
    lines.append(f"TOTAL ISSUES FOUND: {len(all_issues):,}")
    lines.append(f"  Errors:   {issue_severity_counts.get('error', 0):,}")
    lines.append(f"  Warnings: {issue_severity_counts.get('warning', 0):,}")
    lines.append("")
    lines.append("ISSUES BY TYPE (top 15):")
    for issue_type, cnt in issue_type_counts.most_common(15):
        lines.append(f"  {issue_type}: {cnt:,}")
    lines.append("")
    lines.append("=" * 70)

    return "\n".join(lines)


def _format_size(size_bytes):
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} PB"
